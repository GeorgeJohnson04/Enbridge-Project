"""
Fits the four log-OLS oil price models and writes results to
output/OLS_Model_Results.xlsx.

Segmentation:
    Light crude  (B1=1) -> dependent variable = log(WTI_real)
    Heavy crude  (B1=0) -> dependent variable = log(WCS_real)
    Short-term  (B2=0) -> rows where war_long == 0 (peace + flare-ups <6mo)
    Long-term   (B2=1) -> rows where war_long == 1 (sustained conflict >=6mo)

Four models => 2 (crude) x 2 (war state) cells.
"""
from __future__ import annotations

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import config

# (crude_label, B1, dep_var, lag_var)
CRUDE_SPECS = [
    ("Light", 1, "log_wti_real", "log_wti_lag1"),
    ("Heavy", 0, "log_wcs_real", "log_wcs_lag1"),
]
# (war_label, B2)
WAR_SPECS = [
    ("Short-term", 0),
    ("Long-term",  1),
]

# Map config.REGRESSORS placeholder name "log_price_lag1" to the per-crude variant
def _resolve_regressors(lag_var: str) -> list[str]:
    return [lag_var if x == "log_price_lag1" else x for x in config.REGRESSORS]


def _fit_one(df: pd.DataFrame, dep: str, regressors: list[str]) -> sm.regression.linear_model.RegressionResultsWrapper:
    cols = [dep] + regressors
    sub = df[cols].dropna()
    if len(sub) < len(regressors) + 5:
        raise ValueError(f"Too few observations: {len(sub)} (need {len(regressors)+5}+)")
    X = sm.add_constant(sub[regressors])
    y = sub[dep]
    return sm.OLS(y, X).fit()


def _coef_table(res) -> pd.DataFrame:
    return pd.DataFrame({
        "coef":      res.params,
        "std_err":   res.bse,
        "t":         res.tvalues,
        "p":         res.pvalues,
        "ci_low":    res.conf_int()[0],
        "ci_high":   res.conf_int()[1],
    }).round(5)


def _diag_table(res, n_obs: int) -> pd.DataFrame:
    return pd.DataFrame({
        "metric": ["n_obs", "R2", "adj_R2", "F-stat", "F p-value", "AIC", "BIC", "Durbin-Watson"],
        "value":  [n_obs, res.rsquared, res.rsquared_adj, res.fvalue, res.f_pvalue,
                   res.aic, res.bic, sm.stats.stattools.durbin_watson(res.resid)],
    }).round(4)


def _write_sheet(wb: Workbook, name: str, header: dict, coef_df: pd.DataFrame,
                 diag_df: pd.DataFrame, resid_df: pd.DataFrame | None) -> None:
    ws = wb.create_sheet(name[:31])
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DCE6F1")

    ws.append([f"Model: {name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Specification"])
    ws["A3"].font = bold
    for k, v in header.items():
        ws.append([k, v])
    ws.append([])

    ws.append(["Coefficients"])
    ws.cell(row=ws.max_row, column=1).font = bold
    coef_df = coef_df.reset_index().rename(columns={"index": "variable"})
    for r in dataframe_to_rows(coef_df, index=False, header=True):
        ws.append(r)
    for cell in ws[ws.max_row - len(coef_df)]:
        cell.font = bold
        cell.fill = hdr_fill
    ws.append([])

    ws.append(["Diagnostics"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for r in dataframe_to_rows(diag_df, index=False, header=True):
        ws.append(r)
    for cell in ws[ws.max_row - len(diag_df)]:
        cell.font = bold
        cell.fill = hdr_fill

    if resid_df is not None and len(resid_df):
        ws.append([])
        ws.append(["Residuals (sample)"])
        ws.cell(row=ws.max_row, column=1).font = bold
        for r in dataframe_to_rows(resid_df.tail(24).reset_index(), index=False, header=True):
            ws.append(r)

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 28)


def run_all() -> None:
    df = pd.read_csv(config.FEATURES_CSV, index_col="date", parse_dates=True)

    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet collected as we go
    summary_rows = []

    for crude_label, b1, dep, lag in CRUDE_SPECS:
        for war_label, b2 in WAR_SPECS:
            sub = df[df["war_long"] == b2].copy()
            regressors = _resolve_regressors(lag)
            try:
                res = _fit_one(sub, dep, regressors)
            except ValueError as e:
                print(f"  [SKIP] {crude_label} | {war_label}: {e}")
                continue

            n_obs = int(res.nobs)
            sheet_name = f"{crude_label}_{war_label.replace('-', '')}"
            header = {
                "Crude type (B1)":   f"{crude_label} ({b1})",
                "War state (B2)":    f"{war_label} ({b2})",
                "Dependent variable": dep,
                "Sample size":        n_obs,
                "Sample period":      f"{sub.dropna(subset=[dep]+regressors).index.min().date()} to {sub.dropna(subset=[dep]+regressors).index.max().date()}",
            }
            coef = _coef_table(res)
            diag = _diag_table(res, n_obs)
            resid = pd.DataFrame({"actual": res.fittedvalues + res.resid,
                                  "fitted": res.fittedvalues,
                                  "residual": res.resid})
            _write_sheet(wb, sheet_name, header, coef, diag, resid)

            summary_rows.append({
                "Model":     sheet_name,
                "Crude":     crude_label,
                "War state": war_label,
                "n":         n_obs,
                "R2":        round(res.rsquared, 4),
                "adj_R2":    round(res.rsquared_adj, 4),
                "F p-value": round(res.f_pvalue, 4),
                "DW":        round(sm.stats.stattools.durbin_watson(res.resid), 3),
            })
            print(f"  [OK]   {crude_label:5s} | {war_label:10s}  n={n_obs:3d}  R2={res.rsquared:.3f}  adj_R2={res.rsquared_adj:.3f}")

    # Summary sheet
    summary = pd.DataFrame(summary_rows)
    ws = wb.create_sheet("Summary", 0)
    ws.append(["OLS Oil Price Models - Summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Notes"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    notes = [
        "Dependent variable is log of inflation-adjusted price (real USD).",
        f"Real prices in {config.REAL_PRICE_BASE_YEAR} dollars; CPI = FRED CPIAUCSL.",
        f"Long-term war = active conflict lasting >= {config.LONG_WAR_MONTHS} months.",
        "Short-term segment includes peacetime months (no active long war).",
        "Light crude = WTI; Heavy crude = WCS (user CSV preferred, else WTI - $15 proxy).",
        "B1 (crude) and B2 (war) act as model selectors, not regressors.",
    ]
    for n in notes:
        ws.append([n])
    ws.append([])
    ws.append(["Model fit summary"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws.append(r)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col), default=10
        ) + 2

    # Data sheet (last 60 months for reference, full file is data/panel_features.csv)
    ws = wb.create_sheet("Data_recent")
    recent = df.tail(60).reset_index()
    for r in dataframe_to_rows(recent, index=False, header=True):
        ws.append(r)

    wb.save(config.RESULTS_XLSX)
    print(f"\nResults saved: {config.RESULTS_XLSX}")


if __name__ == "__main__":
    run_all()
